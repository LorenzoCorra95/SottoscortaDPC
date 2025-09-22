import streamlit as st
import pandas as pd
import numpy as np
import openpyxl as op
import io
import datetime as dt

st.set_page_config(page_title="Analisi Sottoscorta DPC", layout="wide")
st.title("📊 Analisi Sottoscorta - DPC")

st.write("Carica tutti i file CSV richiesti:")

# --- File uploader ---
file_contratti = st.file_uploader("Contratti", type=[".csv"])
file_ordini = st.file_uploader("Ordini", type=[".csv"])
file_anag = st.file_uploader("Anagrafica", type=[".csv"])
file_carichi = st.file_uploader("Carichi", type=[".csv"])
file_sottoscorta = st.file_uploader("Sottoscorte/Fabbisogni", type=[".csv"])
file_carenze = st.file_uploader("Carenze", type=[".csv"])

numero_gg_riordino=st.selectbox("Selezionare autonomia sotto la quale riordinare",[i for i in range(50,76)])

if st.button("Esegui analisi"):
    if not all([file_contratti, file_ordini, file_anag, file_carichi, file_sottoscorta, file_carenze, numero_gg_riordino]):
        st.error("⚠️ Devi caricare tutti i file CSV e indicare l'autonomia!")
    else:
        # --- Creazione DataFrame ---
        df_c = pd.read_csv(file_contratti, sep=";", encoding="latin")
        df_o = pd.read_csv(file_ordini, sep=";", encoding="latin",dtype={"Cod.Prodotto/Fornitore":str})
        df_anag = pd.read_csv(file_anag, sep=";", encoding="latin", dtype={"Minsan": str})
        df_carichi = pd.read_csv(file_carichi, sep=";", encoding="latin")
        df_sott = pd.read_csv(file_sottoscorta, sep=";", encoding="latin", dtype={"Minsan": str})
        df_carenze = pd.read_csv(file_carenze, sep=";", encoding="latin", dtype={"Minsan": str})

        # --- Prodotti da escludere ---
        prodEscl=[
        "042494070","042494029","044924025","043208091","043208038","045183050","045183148","043443136","043443047","044229312","044229223","044229045","044229134",
        "043145022","043145061","043375118","043375082","043375056","043375029","046343113","046343253","046339089","046339026","046342085","046342022"
    ]

        # ----------------------------------------------------------------------------------------
        # Sezione formati df
        
        # sistemo il df dell'anagrafica
        df_anag["Scadenza"].unique()
        df_anag=df_anag[(df_anag["Scadenza"].isna()==False)&(df_anag["Scadenza"]!='01/01/1900')]
        
        # sistemo il formato del df contratti
        df_c=df_c.iloc[:,[0,1,2,3,4,5,6,8,9,10,11,22,23,24,26,27,37,38]]
        
        for data in ["Validità dal","Validità al"]:
            df_c[data]=pd.to_datetime(df_c[data])
        for val in ["Importo","Prezzo"]:
            df_c[val]=df_c[val].apply(lambda x:float(x.replace(",",".")))
        for val in ["Qta","Qta Ordinato"]:
            try:
                df_c[val]=df_c[val].astype(int)
            except:
                df_c[val]=df_c[val].apply(lambda x: int(x[0:x.find(",")]) if x.find(",")>0 else int(x))
        
        Colonne={"Tipo contratto":"TipoContratto",
                 "Validità dal":"DataIn",
                 "Validità al":"DataFin",
                 "Stato":"StatoContratto",
                 "Codice CIG":"CIG",
                 "Descrizione CIG":"DescrizioneCIG",
                 "Stato Riga":"StatoRiga",
                 "Prodotto":"CodProd",
                 "Descr.":"Prodotto",
                 "Qta Ordinato":"QtaOrdinato"}
        
        df_c.rename(columns=Colonne,inplace=True)
        df_c.insert(13,"Minsan",pd.merge(df_c["CodProd"],df_anag[["CodAreas","Minsan"]],left_on="CodProd",right_on="CodAreas",how="left")["Minsan"])
        df_c.insert(18,"QtaResidua",df_c["Qta"]-df_c["QtaOrdinato"])
        
        # sistemo il formato del df ordini
        df_o.insert(0,"Ordine","DPC-"+df_o["Anno"].astype(str)+"-"+df_o["Num."].astype(str))
        df_o["Data ordine"]=pd.to_datetime(df_o["Data ordine"])
        df_o=df_o.iloc[:,[0,6,8,9,14,15,16,19,22,42]]
        df_o.rename(columns={"Data ordine":"Data",
                             "Qta/Val Rettificata":"Qta",
                             "Prezzo Unit.":"Prezzo",
                             "Cod.Prodotto/Fornitore":"Minsan"},inplace=True)
        
        
        #sistemo il formato del df carichi
        df_carichi["Data Attività"]=pd.to_datetime(df_carichi["Data Attività"].str.slice(0,10),format="%d/%m/%Y")
        df_carichi["Minsan"]=df_carichi["Prodotto"].str.slice(0,9)
        df_carichi["Prodotto"]=df_carichi["Prodotto"].str.slice(12)
        df_carichi=df_carichi[(df_carichi["Riferimento Ordine Carico"].str.contains("DPC24")==False)&
                              (df_carichi["Riferimento Ordine Carico"].str.len() == 9)]
        
        df_carichi=df_carichi.iloc[:,[0,1,5,2,3,4]].rename(columns={"Data Attività":"Data","Riferimento Ordine Carico":"Ordine",
                                                                    "Qta Movimentata":"QtaCaricata"})
        df_carichi["Ordine"]="DPC-2025-" + df_carichi["Ordine"].str.slice(5).astype(int).astype(str)
        
        # sistemo il formato del df sottoscorta
        df_sott=df_sott.fillna("")
        
        
        indici=[]
        for i in range(2):
            lista=df_sott.iloc[i].tolist()
            indici.append(lista)
        
        new_indice=[]
        for i in range(len(lista)):
            val1=str(indici[0][i])
            val2=str(indici[1][i])
            if val1=="":
                new_indice.append(val2)
            elif val2=="":
                new_indice.append(val1)
            else:
                new_indice.append(val1+"-"+val2)
        
        df_sott.columns=new_indice
        df_sott.drop(axis=0, index=[0,1], inplace=True)
        df_sott.replace("", np.nan, inplace=True)
        df_sott.dropna(axis=1, how='all', inplace=True)
        
        df_sott.info()
        df_sott=df_sott.iloc[:,[0,1,2,4,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,22,23,28]]
        df_sott.rename(columns={
            "Domanda Media Giornaliera":"Cmg",
            "Giacenza Totale":"Giacenza"},inplace=True)
        
        for val in ["Cmg","Giacenza"]:
            df_sott[val]=df_sott[val].apply(lambda x: float(x.replace(",",".")))
        
        # ----------------------------------------------------------------------------------------
        # Sezione elaborazione df
        
        # ORDINI
        
        df_o=df_o[(df_o["Stato"]!="Revocato")&(df_o["Qta"]!=0)] # elimino gli ordini revocati e le righe con qta=0
        
        carichiOrd=df_carichi.groupby(["Minsan","Ordine"])["QtaCaricata"].sum().reset_index() # calcolo la qta caricata per minsan-ordine
        
        # aggiungo al df degli ordini l'informazione della qta caricata e di quella ancora da caricare
        df_o = df_o.merge(carichiOrd, on=["Minsan", "Ordine"], how="left").reset_index()

        df_o["QtaCaricata"] = df_o["QtaCaricata"].fillna(0)
        df_o["DaCaricare"]=df_o["Qta"]-df_o["QtaCaricata"]
        
        OrdiniPendenti=df_o[df_o["DaCaricare"]>0] # ordini con carico parziale/non evasi
        DaCaricare=OrdiniPendenti.groupby("Minsan")["DaCaricare"].sum().reset_index() # per ogni minsan calcolo la qta ancora da caricare
        
        # per ogni minsan individuo l'ultimo ordine ancora aperto
        UltimoOrdPendente=OrdiniPendenti[["Minsan","Ordine","Data","Qta"]].sort_values(by=["Minsan","Data"],ascending=[True,False]).drop_duplicates(subset="Minsan")
        
        ProdDanno=df_o[df_o["Autorizzazione"]=="DPC/2025/1-2"][["Minsan","Fornitore"]].drop_duplicates() #ordini in danno
        
        
        # CONTRATTI
        
        # Tipo Contratti -> ['CONTRATTO DPC IN GARA', 'CONTRATTI DPC IN ECONOMIA']
        # Filtro il df dei contratti per i soli minsan presenti in anagrafica
        df_c=df_c.loc[df_c["Minsan"].isin(df_anag["Minsan"])]
        ContAperti=df_c[(df_c["StatoRiga"]=="Aperto")&(df_c["StatoContratto"]=="Aperto")] # contratti aperti con riga aperta
        DispProd=ContAperti.groupby("Minsan")["QtaResidua"].sum().reset_index() # per i contratti aperti con riga aperta calcolo la somma della qta disp per minsan
        
        
        #SOTTOSCORTA
        
        GruppoEq=df_sott.groupby("GruppoEq")[["Cmg","Giacenza"]].sum().reset_index() # per ogni gruppo eq calcolo la somma di cmg e giacenza
        
        
        # eseguo tutti i merge necessari per recuperare le info dagli altri df
        df_sott=pd.merge(
            df_sott,
            GruppoEq,
            on="GruppoEq",
            suffixes=["Prod","GruppoEq"],
            how="left")
        
        df_sott=pd.merge(
            df_sott,
            df_anag[["Minsan", "Pa","Fornitore","Frigo","TipoAcq","Gara"]],
            on="Minsan",
            how="left")
        
        df_sott=pd.merge(
            df_sott,
            DispProd,
            on="Minsan",
            how="left")
        df_sott["QtaResidua"]=df_sott["QtaResidua"].fillna(0)
        
        
        df_sott=pd.merge(
            df_sott,
            DaCaricare,
            on="Minsan",
            how="left")
        
        df_sott["DaCaricare"]=df_sott["DaCaricare"].fillna(0)
        
        df_sott=pd.merge(
            df_sott,
            UltimoOrdPendente[["Minsan","Ordine","Data"]],
            on="Minsan",
            how="left"
            )
        
        df_sott=pd.merge(
            df_sott,
            df_carenze,
            on="Minsan",
            how="left")
        
        
        # per i record in cui il gruppo eq è nullo riporto il cmg e la giacenza = a quello del prodotto (e non = al gruppo eq)
        df_sott.loc[df_sott["GruppoEq"].isnull(),"CmgGruppoEq"]=df_sott["CmgProd"]
        df_sott.loc[df_sott["GruppoEq"].isnull(),"GiacenzaGruppoEq"]=df_sott["GiacenzaProd"] 
        
        # creo le categorie prodotto da assegnare a tutti i minsan non presenti in anagrafica
        CategorieProdotto=[
            df_sott.loc[~df_sott["Minsan"].isin(df_anag["Minsan"])]["Minsan"].isin(prodEscl),
            df_sott.loc[~df_sott["Minsan"].isin(df_anag["Minsan"])]["Minsan"].isin(ProdDanno["Minsan"])
            ]
        
        Cat1=["IN ESAURIMENTO","DANNO"]
        
        # assegno le categorie ai prodotti non presenti in anagrafica
        df_sott.loc[~df_sott["Minsan"].isin(df_anag["Minsan"]),"TipoAcq"]=np.select(CategorieProdotto,Cat1,default="non classificato")
        
        
        # per i record in cui il tipo prodotto è AQ riporto il cmg e la giacenza = a quello del prodotto (e non = al gruppo eq)
        df_sott.loc[(df_sott["TipoAcq"]=="AQ_1")|(df_sott["TipoAcq"]=="AQ_2"),"CmgGruppoEq"]=df_sott["CmgProd"]
        df_sott.loc[(df_sott["TipoAcq"]=="AQ_1")|(df_sott["TipoAcq"]=="AQ_2"),"GiacenzaGruppoEq"]=df_sott["GiacenzaProd"]
        
        
        # per ogni prodotto calcolo l'attuale autonomia
        df_sott["Autonomia"]=(df_sott["GiacenzaGruppoEq"]+df_sott["DaCaricare"])/df_sott["CmgGruppoEq"]
        df_sott["Autonomia"]=df_sott["Autonomia"].fillna(9999)
        df_sott.loc[df_sott["Autonomia"]==np.inf, "Autonomia"]=9999
        df_sott["Autonomia"]=df_sott["Autonomia"].astype(int)
        
        # per i prodotti acquistati in danno recupero l'info del fornitore
        df_sott.loc[df_sott["TipoAcq"]=="DANNO","Fornitore"]=pd.merge(
            df_sott.loc[df_sott["TipoAcq"]=="DANNO"],
            ProdDanno,
            on="Minsan",
            how="left")["Fornitore_y"].values
        
        #funzione per valutare, per ogni fornitore, se almeno un prodotto ha autonomia <= 45
        def OrdineForn(forn):
            df=df_sott[(df_sott["Fornitore"]==forn)&(df_sott["Autonomia"]<=45)]
            return len(df)
        
        #applico la funzione solo ai tipi prodotto da ordinare
        
        df_sott.loc[df_sott["TipoAcq"].isin(["GARA","DANNO","ECONOMIA","AQ"]),"QtaDaOrdinare"]= (
            df_sott.loc[df_sott["TipoAcq"].isin(["GARA","DANNO","ECONOMIA","AQ"])].apply(lambda x: int((75*x["CmgGruppoEq"]-(x["GiacenzaGruppoEq"]+x["DaCaricare"]))) if
                                              OrdineForn(x["Fornitore"])>=1 and x["Autonomia"]<52 else 0, axis=1)
            )
        
        df_sott["QtaDaOrdinare"]=df_sott["QtaDaOrdinare"].fillna(0)
        
        # seconda categoria prodotto per assegnare lo stato della riga contratto
        CategorieProdotto=[
            ~df_sott["Minsan"].isin(df_c["Minsan"]),
            df_sott["Minsan"].isin(df_c[df_c["StatoContratto"]=="Aperto"]["Minsan"]),
            ~df_sott["Minsan"].isin(df_c[df_c["StatoContratto"]=="Aperto"]["Minsan"])
            ]
        Cat2=["","APERTO","CHIUSO"]
        
        df_sott["StatoContratto"]=np.select(CategorieProdotto,Cat2,default="non classificato")
        
        
        df_sott.info()
        
        # ridimensiono il df del sottoscorta e creo il sub-df con l'info dei soli prodotti da oridnare
        intestazioni=["Minsan","Descrizione","Fornitore","Pa","GruppoEq","Frigo","DaCaricare","Ordine","Data","Nota ordine","CmgProd","CmgGruppoEq","GiacenzaProd",
                      "GiacenzaGruppoEq","Autonomia","TipoAcq","Gara","StatoContratto","QtaResidua","QtaDaOrdinare"]
        
        
        indiceInt=[indice for indice in [df_sott.columns.get_loc(i) for i in intestazioni]]
        
        df_sott=df_sott.iloc[:,[5,6,7,8,9,10,11,12,13,14,15,16,17,18]+indiceInt]
        
        df_sott=df_sott.sort_values(by=["Fornitore","Descrizione","Frigo","TipoAcq"])
        
        df_sott_ord=df_sott[(df_sott["QtaDaOrdinare"]>0) & 
                            (~df_sott["TipoAcq"].isin(["IN ESAURIMENTO","FUORI GARA"]))].sort_values(by=["Fornitore","Descrizione","Frigo"])

        # --- Salvataggio Excel in memoria ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as ex:
            df_sott.to_excel(ex, sheet_name="sottoscorta webdpc", index=False)
            df_sott_ord.to_excel(ex, sheet_name="proposta d'ordine", index=False)

        # --- Formattazione con openpyxl ---
        output.seek(0)
        wb = op.load_workbook(output)
        for foglio in wb:
            n_col = foglio.max_column
            foglio.cell(1, n_col+1, value="Nuova Autonomia")
            contaR=1
    
            for riga in foglio:
                contaC=1
                for cella in riga:
                    cella.font=op.styles.Font(name="Calibri", size=12)
                    cella.border=op.styles.borders.Border(
                        top=op.styles.borders.Side(style='thin'),
                        left=op.styles.borders.Side(style='thin'),
                        right=op.styles.borders.Side(style='thin'),
                        bottom=op.styles.borders.Side(style='thin'))
                    cella.alignment=op.styles.Alignment(vertical="center")
                    if contaR==1:
                        cella.font=op.styles.Font(name="Calibri", size=12, bold=True)
                        cella.alignment=op.styles.Alignment(vertical="center",horizontal="center")
                    if contaR!=1 and contaC==n_col+1:
                        cella.value=f'=IFERROR(ROUND((Ai{contaR}+Ac{contaR}+v{contaR})/Y{contaR},0),"")'
                    contaC+=1
                contaR+=1    
    
            foglio.column_dimensions.group("a","n",hidden=True)

        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.success("✅ Analisi completata!")
        st.download_button(
            label="📥 Scarica sottoscorta.xlsx",
            data=final_output,
            file_name="sottoscorta.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )








































